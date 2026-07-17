from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새 시트 기본이름으로 생성
ws.title = "MySheet" # 시트 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # 시트 탭 색상 변경

ws1 = wb.create_sheet("YourSheet") # 새 시트 이름 지정하여 생성 
ws3 = wb.create_sheet("NewSheet", 2) # 새 시트 이름 지정하고 위치 지정하여 생성 

new_ws = wb["NewSheet"] # 시트 이름으로 접근

print(wb.sheetnames) # 모든 시트 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws) # 시트 복사
target.title = "CopiedSheet" # 복사한 시트 이름 변경

wb.save("sample.xlsx")