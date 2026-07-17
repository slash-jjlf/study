# from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # 수식 그대로 가져옴
# for row in ws.values:
#     for cell in row:
#         print(cell)

from openpyxl import load_workbook
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터를 가지고 옴
for row in ws.values:
    for cell in row:
        print(cell)