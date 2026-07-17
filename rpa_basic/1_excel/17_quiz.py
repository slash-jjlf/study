# # [현재까지 작성된 최종 성적 데이터]
# # 학번, 출석, 퀴즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트
# # 1,10,8,5,14,26,12
# # 2,7,3,7,15,24,18
# # 3,9,5,8,8,12,4
# # 4,7,8,7,17,21,18
# # 5,7,8,7,16,25,15
# # 6,3,5,8,8,17,0
# # 7,4,9,10,16,27,18
# # 8,6,6,6,15,19,17
# # 9,10,10,9,19,30,19
# # 10,9,8,8,20,25,20

# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active

# # 데이터 입력하기
# ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])

# scores = [
#     [1,10,8,5,14,26,12],
#     [2,7,3,7,15,24,18],
#     [3,9,5,8,8,12,4],
#     [4,7,8,7,17,21,18],
#     [5,7,8,7,16,25,15],
#     [6,3,5,8,8,17,0],
#     [7,4,9,10,16,27,18],
#     [8,6,6,6,15,19,17],
#     [9,10,10,9,19,30,19],
#     [10,9,8,8,20,25,20]
# ]   

# for s in scores:
#     ws.append(s)

# # 퀴즈2 점수를 모두 10점으로 변경
# for cell in ws["D"]:
#     if cell.row ==1 : # 첫 번째 행은 제외
#         continue
#     cell.value = 10

# # 총점 열을 추가하고, 총점 계산하기
# ws["H1"] = "총점" # H1 셀에 '총점' 입력
# for i in ws["H"]:
#     if i.row == 1: # 첫 번째 행은 제외
#         continue
#     ws[f"H{i.row}"] = f"=SUM(B{i.row}:G{i.row})" # H열에 총점 계산 수식 입력
# wb.save("score.xlsx")

from openpyxl import load_workbook
wb = load_workbook("score.xlsx", data_only=True) # 수식이 아닌 값으로 불러오기
ws = wb.active

# 성적 정보 열을 추가하고, 총점 90점 이상 A, 80점 이상 B, 70점 이상 C, 나머지 D
# 출석이 5미만인 학생은 총점과 상관없이 F 처리
ws['I1'] = "성적" # I1 셀에 '성적' 입력
for i in ws["I"]:
    if i.row == 1: # 첫 번째 행은 제외
        continue
    if ws[f"B{i.row}"].value < 5: # 출석이 5미만이면 F 처리
        ws[f"I{i.row}"] = "F"
    elif ws[f"H{i.row}"].value >= 90:
        ws[f"I{i.row}"] = "A"
    elif ws[f"H{i.row}"].value >= 80:
        ws[f"I{i.row}"] = "B"
    elif ws[f"H{i.row}"].value >= 70:
        ws[f"I{i.row}"] = "C"
    else:
        ws[f"I{i.row}"] = "D"

wb.save("score.xlsx")