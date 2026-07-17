from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

# col_B = ws["B"]
# for cell in col_B:
#     print(cell.value)

# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()






wb.save("sample.xlsx")