from openpyxl import load_workbook
wb = load_workbook("sample_merge.xlsx")
ws = wb.active

# 병합된 셀 해제하기
ws.unmerge_cells("B2:D2") # B2부터 D2까지 병합 해제
wb.save("sample_unmerge.xlsx")
