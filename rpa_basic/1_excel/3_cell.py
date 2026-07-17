from openpyxl import Workbook
wb = Workbook()
ws = wb.active # 기본 시트 활성화
ws.title = "NadoSheet"

# 셀에 값 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

c =ws.cell(row=1, column=3, value=10)
print(c.value)

from random import *

for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=randint(0,100))

wb.save("sample.xlsx")
