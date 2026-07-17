from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 셀 값 가져오기
for x in range(1, 11):
    for y in range(1,11):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()